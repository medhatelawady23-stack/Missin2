"""
Dashboard views - لوحة التحكم الكاملة المستقلة
CRUD لجميع النماذج (بما فيها أنواع العمل والوحدات للمستخدمين) + إدارة التقارير + تصدير PDF/Excel
"""
import io
from datetime import datetime

from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User
from django.contrib import messages
from django import forms
from django.views.generic import (
    TemplateView, ListView, CreateView, UpdateView, DeleteView, DetailView, View
)
from django.urls import reverse_lazy
from django.db.models import Q, Count
from django.http import HttpResponse, FileResponse
from django.shortcuts import get_object_or_404, redirect

from accounts.models import UserProfile
from locations.models import Unit, LocationType, Location
from notes.models import Note, NoteLocation, WorkType, WorkItemDetail, WorkActivity
from .models import Report, ReportItem, ReportItemImage


# ─── Mixin للتحقق من صلاحيات الإدارة ────────────────────────────────────────
class StaffRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    login_url = reverse_lazy('accounts:admin_login')

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(self.request, 'يرجى تسجيل الدخول بحساب إداري للوصول إلى لوحة التحكم')
        return redirect('accounts:admin_login')


# ─── Dashboard Home (Single Unified Control Page) ────────────────────────────
class DashboardHomeView(StaffRequiredMixin, ListView):
    model = Report
    template_name = 'dashboard/home.html'
    context_object_name = 'reports'
    paginate_by = 15

    def get_queryset(self):
        qs = Report.objects.select_related('user', 'unit', 'location', 'location_type').order_by('-created_at')
        q = self.request.GET.get('q')
        unit_id = self.request.GET.get('unit')
        user_id = self.request.GET.get('user')
        location_type_id = self.request.GET.get('location_type')
        location_id = self.request.GET.get('location')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if q:
            qs = qs.filter(
                Q(report_number__icontains=q) |
                Q(location__name__icontains=q) |
                Q(user__username__icontains=q) |
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q)
            )
        if unit_id:
            qs = qs.filter(unit_id=unit_id)
        if user_id:
            qs = qs.filter(user_id=user_id)
        if location_type_id:
            qs = qs.filter(location_type_id=location_type_id)
        if location_id:
            qs = qs.filter(location_id=location_id)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        filtered_qs = self.get_queryset()

        # إحصائيات مؤشرات العمل بناءً على الفلاتر النشطة
        ctx['filtered_reports_count'] = filtered_qs.count()
        
        # عدد التقارير المرفوعة لكل موقع
        ctx['location_breakdown'] = filtered_qs.values(
            'location__id', 'location__name', 'unit__name', 'location_type__name'
        ).annotate(
            report_count=Count('id')
        ).order_by('-report_count')

        # توزيع الأنشطة حسب أنواع العمل
        ctx['work_type_breakdown'] = ReportItem.objects.filter(
            report__in=filtered_qs, work_type__isnull=False
        ).values('work_type__name').annotate(
            item_count=Count('id')
        ).order_by('-item_count')

        # الإحصائيات العامة
        ctx['total_contractors'] = User.objects.filter(Q(profile__role='contractor') | Q(profile__isnull=True)).count()
        ctx['total_supervisors'] = User.objects.filter(profile__role='supervisor').count()
        ctx['total_users'] = User.objects.count()
        ctx['total_units'] = Unit.objects.count()
        ctx['total_locations'] = Location.objects.count()
        ctx['total_location_types'] = LocationType.objects.count()
        ctx['total_notes'] = Note.objects.count()
        ctx['total_work_types'] = WorkType.objects.count()
        ctx['total_reports'] = Report.objects.count()

        # بيانات خيارات الفلاتر (الوحدة -> نوع الموقع -> الموقع)
        ctx['units'] = Unit.objects.filter(is_active=True)
        ctx['users'] = User.objects.filter(is_active=True)
        ctx['location_types'] = LocationType.objects.filter(is_active=True)

        # كاسكاد الفلاتر للمواقع
        unit_id = self.request.GET.get('unit')
        location_type_id = self.request.GET.get('location_type')
        loc_qs = Location.objects.filter(is_active=True)
        if unit_id:
            loc_qs = loc_qs.filter(unit_id=unit_id)
        if location_type_id:
            loc_qs = loc_qs.filter(location_type_id=location_type_id)
        ctx['locations'] = loc_qs

        return ctx


class WorkIndicatorsView(StaffRequiredMixin, ListView):
    model = Report
    template_name = 'dashboard/indicators.html'
    context_object_name = 'reports'
    paginate_by = 20

    def get_queryset(self):
        qs = Report.objects.select_related('user', 'unit', 'location', 'location_type').order_by('-created_at')
        role = self.request.GET.get('role', 'contractor')
        category = self.request.GET.get('category')
        q = self.request.GET.get('q')
        unit_id = self.request.GET.get('unit')
        user_id = self.request.GET.get('user')
        location_type_id = self.request.GET.get('location_type')
        location_id = self.request.GET.get('location')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if role == 'supervisor':
            qs = qs.filter(user__profile__role='supervisor')
        elif role == 'contractor':
            qs = qs.filter(Q(user__profile__role='contractor') | Q(user__profile__isnull=True))

        if category == 'restroom':
            qs = qs.filter(
                Q(location_type__name__icontains='دورة') |
                Q(location_type__name__icontains='مياه') |
                Q(location_type__name__icontains='حمام')
            )
        elif category == 'location':
            qs = qs.filter(
                Q(location_type__name__icontains='موقع') |
                Q(location_type__name__icontains='مواقف') |
                Q(location_type__name__icontains='مبنى')
            )
        elif category == 'open_space':
            qs = qs.filter(
                Q(location_type__name__icontains='انتشار') |
                Q(location_type__name__icontains='ساحة') |
                Q(location_type__name__icontains='حديقة') |
                Q(location_type__name__icontains='شارع') |
                Q(location_type__name__icontains='طريق')
            )

        if q:
            qs = qs.filter(
                Q(report_number__icontains=q) |
                Q(location__name__icontains=q) |
                Q(user__username__icontains=q) |
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q)
            )
        if unit_id:
            qs = qs.filter(unit_id=unit_id)
        if user_id:
            qs = qs.filter(user_id=user_id)
        if location_type_id:
            qs = qs.filter(location_type_id=location_type_id)
        if location_id:
            qs = qs.filter(location_id=location_id)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        filtered_qs = self.get_queryset()

        ctx['filtered_reports_count'] = filtered_qs.count()
        ctx['total_filtered_items'] = ReportItem.objects.filter(report__in=filtered_qs).count()
        ctx['unique_users_count'] = filtered_qs.values('user').distinct().count()

        # حساب الإحصائيات لأنواع الأعمال الستة المحددة بالضبط
        target_work_types = [
            {'name': 'أعمال النظافة', 'keywords': ['نظاف'], 'icon': 'fas fa-broom', 'color': '#0284c7', 'border': '#0284c7', 'bg': '#e0f2fe'},
            {'name': 'الأعمال الزراعية', 'keywords': ['زراع'], 'icon': 'fas fa-seedling', 'color': '#16a34a', 'border': '#16a34a', 'bg': '#dcfce7'},
            {'name': 'الأعمال المدنية', 'keywords': ['مدن'], 'icon': 'fas fa-building', 'color': '#d97706', 'border': '#d97706', 'bg': '#fef3c7'},
            {'name': 'الأعمال الكهربائية', 'keywords': ['كهرب'], 'icon': 'fas fa-bolt', 'color': '#dc2626', 'border': '#dc2626', 'bg': '#fee2e2'},
            {'name': 'الأعمال الميكانيكية', 'keywords': ['ميكانيك'], 'icon': 'fas fa-cogs', 'color': '#7c3aed', 'border': '#7c3aed', 'bg': '#f3e8ff'},
            {'name': 'أعمال الري', 'keywords': ['ري'], 'icon': 'fas fa-droplet', 'color': '#0891b2', 'border': '#0891b2', 'bg': '#cff4fc'},
        ]

        items_qs = ReportItem.objects.filter(report__in=filtered_qs)
        work_type_stats = []
        for wt in target_work_types:
            q_obj = Q()
            for kw in wt['keywords']:
                q_obj |= Q(work_type__name__icontains=kw)
                q_obj |= Q(work_classification__icontains=kw)
                q_obj |= Q(work_item_detail__name__icontains=kw)
                q_obj |= Q(work_activity__name__icontains=kw)
                q_obj |= Q(note__text__icontains=kw)
            q_obj |= Q(work_type__name=wt['name'])

            count = items_qs.filter(q_obj).distinct().count()
            work_type_stats.append({
                'name': wt['name'],
                'count': count,
                'icon': wt['icon'],
                'color': wt['color'],
                'border': wt['border'],
                'bg': wt['bg']
            })
        ctx['work_type_stats'] = work_type_stats

        ctx['location_breakdown'] = filtered_qs.values(
            'location__id', 'location__name', 'unit__name', 'location_type__name'
        ).annotate(
            report_count=Count('id')
        ).order_by('-report_count')

        ctx['work_type_breakdown'] = ReportItem.objects.filter(
            report__in=filtered_qs, work_type__isnull=False
        ).values('work_type__name').annotate(
            item_count=Count('id')
        ).order_by('-item_count')

        ctx['units'] = Unit.objects.filter(is_active=True)
        role = self.request.GET.get('role', 'contractor')
        category = self.request.GET.get('category', '')
        ctx['current_role'] = role
        ctx['current_category'] = category
        if role == 'supervisor':
            ctx['users'] = User.objects.filter(is_active=True, profile__role='supervisor')
        else:
            ctx['users'] = User.objects.filter(is_active=True).filter(Q(profile__role='contractor') | Q(profile__isnull=True))

        loc_types_qs = LocationType.objects.filter(is_active=True)
        if category != 'restroom':
            loc_types_qs = loc_types_qs.exclude(
                Q(name__icontains='دورة') | Q(name__icontains='مياه') | Q(name__icontains='حمام')
            )
        ctx['location_types'] = loc_types_qs
        unit_id = self.request.GET.get('unit')
        location_type_id = self.request.GET.get('location_type')
        loc_qs = Location.objects.filter(is_active=True)
        if unit_id:
            loc_qs = loc_qs.filter(unit_id=unit_id)
        if location_type_id:
            loc_qs = loc_qs.filter(location_type_id=location_type_id)
        ctx['locations'] = loc_qs
        return ctx


# ─── نموذج إدارة المقاولين ──────────────────────────────────────────────
class ContractorUserForm(forms.ModelForm):
    unit = forms.ModelChoiceField(
        queryset=Unit.objects.filter(is_active=True),
        required=True,
        label='الوحدة التابعة للمقاول',
        widget=forms.Select(attrs={'class': 'form-select'})
    )
    password = forms.CharField(
        widget=forms.PasswordInput(attrs={'class': 'form-control', 'placeholder': 'كلمة المرور'}),
        required=False,
        label='كلمة المرور'
    )

    class Meta:
        model = User
        fields = ['username', 'first_name', 'last_name', 'email', 'is_staff', 'is_active']
        widgets = {
            'username': forms.TextInput(attrs={'class': 'form-control'}),
            'first_name': forms.TextInput(attrs={'class': 'form-control'}),
            'last_name': forms.TextInput(attrs={'class': 'form-control'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
            'is_staff': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            profile = getattr(self.instance, 'profile', None)
            if profile:
                self.fields['unit'].initial = profile.unit

    def save(self, commit=True):
        user = super().save(commit=commit)
        password = self.cleaned_data.get('password')
        if password:
            user.set_password(password)
            if commit:
                user.save()

        unit = self.cleaned_data.get('unit')
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.unit = unit
        profile.role = 'contractor'
        profile.save()
        return user


# ─── نموذج إدارة المراقبين ──────────────────────────────────────────────
class SupervisorUserForm(forms.ModelForm):
    password = forms.CharField(
        widget=forms.PasswordInput(attrs={'class': 'form-control', 'placeholder': 'كلمة المرور'}),
        required=False,
        label='كلمة المرور'
    )

    class Meta:
        model = User
        fields = ['username', 'first_name', 'last_name', 'email', 'is_staff', 'is_active']
        widgets = {
            'username': forms.TextInput(attrs={'class': 'form-control'}),
            'first_name': forms.TextInput(attrs={'class': 'form-control'}),
            'last_name': forms.TextInput(attrs={'class': 'form-control'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
            'is_staff': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }

    def save(self, commit=True):
        user = super().save(commit=commit)
        password = self.cleaned_data.get('password')
        if password:
            user.set_password(password)
            if commit:
                user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.unit = None
        profile.role = 'supervisor'
        profile.save()
        return user


# ─── إدارة المقاولين (المستخدمين) ─────────────────────────────────────────────
class UserListView(StaffRequiredMixin, ListView):
    model = User
    template_name = 'dashboard/users/list.html'
    context_object_name = 'users'
    paginate_by = 20

    def get_queryset(self):
        qs = User.objects.filter(Q(profile__role='contractor') | Q(profile__isnull=True)).select_related('profile__unit').order_by('-date_joined')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(Q(username__icontains=q) | Q(email__icontains=q) | Q(first_name__icontains=q))
        return qs


class UserCreateView(StaffRequiredMixin, CreateView):
    model = User
    form_class = ContractorUserForm
    template_name = 'dashboard/users/form.html'
    success_url = reverse_lazy('dashboard:user_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = 'إضافة مقاول جديد'
        ctx['cancel_url'] = reverse_lazy('dashboard:user_list')
        return ctx

    def form_valid(self, form):
        user = form.save()
        messages.success(self.request, f'تم إنشاء حساب المقاول {user.username} بنجاح')
        return redirect(self.success_url)


class UserUpdateView(StaffRequiredMixin, UpdateView):
    model = User
    form_class = ContractorUserForm
    template_name = 'dashboard/users/form.html'
    success_url = reverse_lazy('dashboard:user_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = f'تعديل حساب مقاول: {self.object.username}'
        ctx['cancel_url'] = reverse_lazy('dashboard:user_list')
        return ctx

    def form_valid(self, form):
        user = form.save()
        messages.success(self.request, f'تم تحديث حساب المقاول {user.username} بنجاح')
        return redirect(self.success_url)


class UserDeleteView(StaffRequiredMixin, DeleteView):
    model = User
    template_name = 'dashboard/confirm_delete.html'
    success_url = reverse_lazy('dashboard:user_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف المقاول بنجاح')
        return super().form_valid(form)


# ─── إدارة المراقبين ────────────────────────────────────────────────────────
class SupervisorListView(StaffRequiredMixin, ListView):
    model = User
    template_name = 'dashboard/supervisors/list.html'
    context_object_name = 'supervisors'
    paginate_by = 20

    def get_queryset(self):
        qs = User.objects.filter(profile__role='supervisor').select_related('profile__unit').order_by('-date_joined')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(Q(username__icontains=q) | Q(email__icontains=q) | Q(first_name__icontains=q))
        return qs


class SupervisorCreateView(StaffRequiredMixin, CreateView):
    model = User
    form_class = SupervisorUserForm
    template_name = 'dashboard/users/form.html'
    success_url = reverse_lazy('dashboard:supervisor_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = 'إضافة مراقب جديد'
        ctx['cancel_url'] = reverse_lazy('dashboard:supervisor_list')
        return ctx

    def form_valid(self, form):
        user = form.save()
        messages.success(self.request, f'تم إنشاء حساب المراقب {user.username} بنجاح')
        return redirect(self.success_url)


class SupervisorUpdateView(StaffRequiredMixin, UpdateView):
    model = User
    form_class = SupervisorUserForm
    template_name = 'dashboard/users/form.html'
    success_url = reverse_lazy('dashboard:supervisor_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = f'تعديل حساب مراقب: {self.object.username}'
        ctx['cancel_url'] = reverse_lazy('dashboard:supervisor_list')
        return ctx

    def form_valid(self, form):
        user = form.save()
        messages.success(self.request, f'تم تحديث حساب المراقب {user.username} بنجاح')
        return redirect(self.success_url)


class SupervisorDeleteView(StaffRequiredMixin, DeleteView):
    model = User
    template_name = 'dashboard/confirm_delete.html'
    success_url = reverse_lazy('dashboard:supervisor_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف المراقب بنجاح')
        return super().form_valid(form)


# ─── إدارة الوحدات ───────────────────────────────────────────────────────────
class UnitListView(StaffRequiredMixin, ListView):
    model = Unit
    template_name = 'dashboard/units/list.html'
    context_object_name = 'units'
    paginate_by = 20

    def get_queryset(self):
        qs = Unit.objects.annotate(location_count=Count('locations')).order_by('name')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(name__icontains=q)
        return qs


class UnitCreateView(StaffRequiredMixin, CreateView):
    model = Unit
    template_name = 'dashboard/units/form.html'
    fields = ['name', 'description', 'is_active']
    success_url = reverse_lazy('dashboard:unit_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم إضافة الوحدة بنجاح')
        return super().form_valid(form)


class UnitUpdateView(StaffRequiredMixin, UpdateView):
    model = Unit
    template_name = 'dashboard/units/form.html'
    fields = ['name', 'description', 'is_active']
    success_url = reverse_lazy('dashboard:unit_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث الوحدة بنجاح')
        return super().form_valid(form)


class UnitDeleteView(StaffRequiredMixin, DeleteView):
    model = Unit
    template_name = 'dashboard/confirm_delete.html'
    success_url = reverse_lazy('dashboard:unit_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف الوحدة بنجاح')
        return super().form_valid(form)


# ─── إدارة أنواع المواقع ──────────────────────────────────────────────────────
class LocationTypeListView(StaffRequiredMixin, ListView):
    model = LocationType
    template_name = 'dashboard/location_types/list.html'
    context_object_name = 'location_types'
    paginate_by = 20

    def get_queryset(self):
        qs = LocationType.objects.all().order_by('name')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(name__icontains=q)
        return qs


class LocationTypeCreateView(StaffRequiredMixin, CreateView):
    model = LocationType
    template_name = 'dashboard/location_types/form.html'
    fields = ['name', 'description', 'is_active']
    success_url = reverse_lazy('dashboard:locationtype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم إضافة نوع الموقع بنجاح')
        return super().form_valid(form)


class LocationTypeUpdateView(StaffRequiredMixin, UpdateView):
    model = LocationType
    template_name = 'dashboard/location_types/form.html'
    fields = ['name', 'description', 'is_active']
    success_url = reverse_lazy('dashboard:locationtype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث نوع الموقع بنجاح')
        return super().form_valid(form)


class LocationTypeDeleteView(StaffRequiredMixin, DeleteView):
    model = LocationType
    template_name = 'dashboard/confirm_delete.html'
    success_url = reverse_lazy('dashboard:locationtype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف نوع الموقع بنجاح')
        return super().form_valid(form)


# ─── إدارة أنواع العمل ────────────────────────────────────────────────────────
class WorkTypeListView(StaffRequiredMixin, ListView):
    model = WorkType
    template_name = 'dashboard/work_types/list.html'
    context_object_name = 'work_types'
    paginate_by = 20

    def get_queryset(self):
        qs = WorkType.objects.prefetch_related('details__activities').order_by('name')
        q = self.request.GET.get('q')
        cat = self.request.GET.get('category')
        if q:
            qs = qs.filter(name__icontains=q)
        if cat:
            qs = qs.filter(category=cat)
        return qs


class WorkTypeCreateView(StaffRequiredMixin, CreateView):
    model = WorkType
    template_name = 'dashboard/work_types/form.html'
    fields = ['name', 'category', 'is_active']
    success_url = reverse_lazy('dashboard:worktype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم إضافة نوع العمل بنجاح')
        return super().form_valid(form)


class WorkTypeUpdateView(StaffRequiredMixin, UpdateView):
    model = WorkType
    template_name = 'dashboard/work_types/form.html'
    fields = ['name', 'category', 'is_active']
    success_url = reverse_lazy('dashboard:worktype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث نوع العمل بنجاح')
        return super().form_valid(form)


class WorkTypeDeleteView(StaffRequiredMixin, DeleteView):
    model = WorkType
    template_name = 'dashboard/confirm_delete.html'
    success_url = reverse_lazy('dashboard:worktype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف نوع العمل بنجاح')
        return super().form_valid(form)


# ─── إدارة البنود التفصيلية ───────────────────────────────────────────────────
class WorkItemDetailListView(StaffRequiredMixin, ListView):
    model = WorkItemDetail
    template_name = 'dashboard/work_items/list.html'
    context_object_name = 'work_items'
    paginate_by = 20

    def get_queryset(self):
        qs = WorkItemDetail.objects.select_related('work_type').order_by('work_type__name', 'name')
        q = self.request.GET.get('q')
        work_type_id = self.request.GET.get('work_type')
        if q:
            qs = qs.filter(name__icontains=q)
        if work_type_id:
            qs = qs.filter(work_type_id=work_type_id)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['work_types'] = WorkType.objects.all().order_by('name')
        return ctx


class WorkItemDetailCreateView(StaffRequiredMixin, CreateView):
    model = WorkItemDetail
    template_name = 'dashboard/work_items/form.html'
    fields = ['work_type', 'name', 'is_active']
    success_url = reverse_lazy('dashboard:worktype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم إضافة البند التفصيلي بنجاح')
        return super().form_valid(form)


class WorkItemDetailUpdateView(StaffRequiredMixin, UpdateView):
    model = WorkItemDetail
    template_name = 'dashboard/work_items/form.html'
    fields = ['work_type', 'name', 'is_active']
    success_url = reverse_lazy('dashboard:worktype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث البند التفصيلي بنجاح')
        return super().form_valid(form)


class WorkItemDetailDeleteView(StaffRequiredMixin, DeleteView):
    model = WorkItemDetail
    template_name = 'dashboard/confirm_delete.html'
    success_url = reverse_lazy('dashboard:worktype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف العنصر بنجاح')
        return super().form_valid(form)


# ─── إدارة الأنشطة ─────────────────────────────────────────────────────────────
class WorkActivityCreateView(StaffRequiredMixin, CreateView):
    model = WorkActivity
    template_name = 'dashboard/work_activities/form.html'
    fields = ['element', 'name', 'is_active']
    success_url = reverse_lazy('dashboard:worktype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم إضافة النشاط بنجاح')
        return super().form_valid(form)


class WorkActivityUpdateView(StaffRequiredMixin, UpdateView):
    model = WorkActivity
    template_name = 'dashboard/work_activities/form.html'
    fields = ['element', 'name', 'is_active']
    success_url = reverse_lazy('dashboard:worktype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث النشاط بنجاح')
        return super().form_valid(form)


class WorkActivityDeleteView(StaffRequiredMixin, DeleteView):
    model = WorkActivity
    template_name = 'dashboard/confirm_delete.html'
    success_url = reverse_lazy('dashboard:worktype_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف النشاط بنجاح')
        return super().form_valid(form)


# ─── إدارة المواقع ───────────────────────────────────────────────────────────
class LocationListView(StaffRequiredMixin, ListView):
    model = Location
    template_name = 'dashboard/locations/list.html'
    context_object_name = 'locations'
    paginate_by = 20

    def get_queryset(self):
        qs = Location.objects.select_related('unit', 'location_type').order_by('unit', 'name')
        q = self.request.GET.get('q')
        unit_id = self.request.GET.get('unit')
        type_id = self.request.GET.get('type')
        if q:
            qs = qs.filter(name__icontains=q)
        if unit_id:
            qs = qs.filter(unit_id=unit_id)
        if type_id:
            qs = qs.filter(location_type_id=type_id)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['units'] = Unit.objects.all()
        ctx['location_types'] = LocationType.objects.all()
        return ctx



class LocationCreateView(StaffRequiredMixin, CreateView):
    model = Location
    template_name = 'dashboard/locations/form.html'
    fields = ['unit', 'location_type', 'name', 'is_active']
    success_url = reverse_lazy('dashboard:location_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم إضافة الموقع بنجاح')
        return super().form_valid(form)


class LocationUpdateView(StaffRequiredMixin, UpdateView):
    model = Location
    template_name = 'dashboard/locations/form.html'
    fields = ['unit', 'location_type', 'name', 'is_active']
    success_url = reverse_lazy('dashboard:location_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث الموقع بنجاح')
        return super().form_valid(form)



class LocationDeleteView(StaffRequiredMixin, DeleteView):
    model = Location
    template_name = 'dashboard/confirm_delete.html'
    success_url = reverse_lazy('dashboard:location_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف الموقع بنجاح')
        return super().form_valid(form)


# ─── إدارة الملاحظات ─────────────────────────────────────────────────────────
class NoteListView(StaffRequiredMixin, ListView):
    model = Note
    template_name = 'dashboard/notes/list.html'
    context_object_name = 'notes'
    paginate_by = 20

    def get_queryset(self):
        qs = Note.objects.prefetch_related('locations').order_by('text')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(text__icontains=q)
        return qs


class NoteCreateView(StaffRequiredMixin, CreateView):
    model = Note
    template_name = 'dashboard/notes/form.html'
    fields = ['text', 'is_active']
    success_url = reverse_lazy('dashboard:note_list')

    def form_valid(self, form):
        note = form.save()
        location_ids = self.request.POST.getlist('locations')
        NoteLocation.objects.filter(note=note).delete()
        for loc_id in location_ids:
            NoteLocation.objects.get_or_create(note=note, location_id=loc_id)
        messages.success(self.request, 'تم إضافة الملاحظة بنجاح')
        return redirect(self.success_url)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['all_locations'] = Location.objects.select_related('unit').order_by('unit', 'name')
        return ctx


class NoteUpdateView(StaffRequiredMixin, UpdateView):
    model = Note
    template_name = 'dashboard/notes/form.html'
    fields = ['text', 'is_active']
    success_url = reverse_lazy('dashboard:note_list')

    def form_valid(self, form):
        note = form.save()
        location_ids = self.request.POST.getlist('locations')
        NoteLocation.objects.filter(note=note).delete()
        for loc_id in location_ids:
            NoteLocation.objects.get_or_create(note=note, location_id=loc_id)
        messages.success(self.request, 'تم تحديث الملاحظة بنجاح')
        return redirect(self.success_url)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['all_locations'] = Location.objects.select_related('unit').order_by('unit', 'name')
        ctx['selected_locations'] = list(
            self.object.locations.values_list('id', flat=True)
        )
        return ctx


class NoteDeleteView(StaffRequiredMixin, DeleteView):
    model = Note
    template_name = 'dashboard/confirm_delete.html'
    success_url = reverse_lazy('dashboard:note_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف الملاحظة بنجاح')
        return super().form_valid(form)


# ─── إدارة التقارير ──────────────────────────────────────────────────────────
class ReportListView(StaffRequiredMixin, ListView):
    model = Report
    template_name = 'dashboard/reports/list.html'
    context_object_name = 'reports'
    paginate_by = 20

    def get_queryset(self):
        qs = Report.objects.select_related('user', 'unit', 'location_type', 'location').order_by('-created_at')
        role = self.request.GET.get('role', 'contractor')
        category = self.request.GET.get('category')
        q = self.request.GET.get('q')
        unit_id = self.request.GET.get('unit')
        user_id = self.request.GET.get('user')
        location_type_id = self.request.GET.get('location_type')
        location_id = self.request.GET.get('location')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if role == 'supervisor':
            qs = qs.filter(user__profile__role='supervisor')
        elif role == 'contractor':
            qs = qs.filter(Q(user__profile__role='contractor') | Q(user__profile__isnull=True))

        if category == 'restroom':
            qs = qs.filter(
                Q(location_type__name__icontains='دورة') |
                Q(location_type__name__icontains='مياه') |
                Q(location_type__name__icontains='حمام')
            )
        elif category == 'location':
            qs = qs.filter(
                Q(location_type__name__icontains='موقع') |
                Q(location_type__name__icontains='مواقف') |
                Q(location_type__name__icontains='مبنى')
            )
        elif category == 'open_space':
            qs = qs.filter(
                Q(location_type__name__icontains='انتشار') |
                Q(location_type__name__icontains='ساحة') |
                Q(location_type__name__icontains='حديقة') |
                Q(location_type__name__icontains='شارع') |
                Q(location_type__name__icontains='طريق')
            )

        if q:
            qs = qs.filter(
                Q(report_number__icontains=q) |
                Q(location__name__icontains=q) |
                Q(user__username__icontains=q) |
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q)
            )
        if unit_id:
            qs = qs.filter(unit_id=unit_id)
        if user_id:
            qs = qs.filter(user_id=user_id)
        if location_type_id:
            qs = qs.filter(location_type_id=location_type_id)
        if location_id:
            qs = qs.filter(location_id=location_id)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        filtered_qs = self.get_queryset()

        # إحصائيات مؤشرات العمل بناءً على الفلاتر النشطة
        ctx['filtered_reports_count'] = filtered_qs.count()
        ctx['location_breakdown'] = filtered_qs.values(
            'location__id', 'location__name', 'unit__name', 'location_type__name'
        ).annotate(
            report_count=Count('id')
        ).order_by('-report_count')

        ctx['units'] = Unit.objects.filter(is_active=True)
        role = self.request.GET.get('role', 'contractor')
        category = self.request.GET.get('category', '')
        ctx['current_role'] = role
        ctx['current_category'] = category

        if category == 'restroom':
            ctx['restroom_grouped_reports'] = build_restroom_grouped_data(filtered_qs)
            ctx['restroom_activities'] = RESTROOM_ACTIVITIES

        if role == 'supervisor':
            ctx['users'] = User.objects.filter(is_active=True, profile__role='supervisor')
        else:
            ctx['users'] = User.objects.filter(is_active=True).filter(Q(profile__role='contractor') | Q(profile__isnull=True))

        loc_types_qs = LocationType.objects.filter(is_active=True)
        if category != 'restroom':
            loc_types_qs = loc_types_qs.exclude(
                Q(name__icontains='دورة') | Q(name__icontains='مياه') | Q(name__icontains='حمام')
            )
        ctx['location_types'] = loc_types_qs
        unit_id = self.request.GET.get('unit')
        location_type_id = self.request.GET.get('location_type')
        loc_qs = Location.objects.filter(is_active=True)
        if unit_id:
            loc_qs = loc_qs.filter(unit_id=unit_id)
        if location_type_id:
            loc_qs = loc_qs.filter(location_type_id=location_type_id)
        ctx['locations'] = loc_qs
        return ctx




class ReportDetailView(StaffRequiredMixin, DetailView):
    model = Report
    template_name = 'dashboard/reports/detail.html'
    context_object_name = 'report'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['items'] = self.object.items.select_related('note', 'work_type').prefetch_related('images')
        return ctx


class ReportDeleteView(StaffRequiredMixin, DeleteView):
    model = Report
    template_name = 'dashboard/confirm_delete.html'

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url and 'delete' not in next_url:
            return next_url
        return reverse_lazy('dashboard:report_list')

    def form_valid(self, form):
        messages.success(self.request, 'تم حذف التقرير بنجاح')
        return super().form_valid(form)


RESTROOM_ACTIVITIES = [
    'أدوات النظافة',
    'مواد النظافة',
    'أعمال النظافة الدورية لدورات المياه',
    'الاعمال المدنية',
    'الأعمال الكهربائية',
    'الأعمال الميكانيكية'
]


def build_restroom_grouped_data(reports_qs):
    grouped_dict = {}
    reports = list(
        reports_qs.select_related('user', 'unit', 'location_type', 'location')
                  .prefetch_related('items__work_activity', 'items__work_type', 'items__work_item_detail')
                  .order_by('created_at')
    )
    for report in reports:
        loc_id = report.location_id if report.location else 0
        rep_date = report.created_at.date()
        key = (loc_id, rep_date)

        if key not in grouped_dict:
            grouped_dict[key] = {
                'date_str': report.created_at.strftime('%Y/%m/%d'),
                'unit_name': report.unit.name if report.unit else 'غير محدد',
                'location_name': report.location.name if report.location else 'غير محدد',
                'visits': []
            }

        item_names = set()
        for item in report.items.all():
            if item.work_activity:
                item_names.add(item.work_activity.name.strip())
            if item.work_classification:
                item_names.add(item.work_classification.strip())

        items_status = []
        for act in RESTROOM_ACTIVITIES:
            is_present = 1 if any(act.strip() in name for name in item_names) else 0
            items_status.append(is_present)

        visit_info = {
            'report_id': report.pk,
            'report_number': report.report_number,
            'user_name': report.user.get_full_name() or report.user.username if report.user else 'غير محدد',
            'time_str': report.created_at.strftime('%I:%M %p'),
            'items': items_status
        }
        grouped_dict[key]['visits'].append(visit_info)

    return list(grouped_dict.values())


# ─── تصدير Excel ─────────────────────────────────────────────────────────────
class ExportExcelView(StaffRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = True

        header_fill = PatternFill('solid', fgColor='1A5276')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(border_style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        reports = Report.objects.select_related(
            'user', 'unit', 'location_type', 'location'
        ).prefetch_related('items__note', 'items__work_type', 'items__work_item_detail', 'items__work_activity', 'items__images').order_by('-created_at')

        role = request.GET.get('role', 'contractor')
        category = request.GET.get('category')
        if role == 'supervisor':
            reports = reports.filter(user__profile__role='supervisor')
        elif role == 'contractor':
            reports = reports.filter(Q(user__profile__role='contractor') | Q(user__profile__isnull=True))

        if category == 'restroom':
            reports = reports.filter(
                Q(location_type__name__icontains='دورة') |
                Q(location_type__name__icontains='مياه') |
                Q(location_type__name__icontains='حمام')
            )
        else:
            # استبعاد دورات المياه وإبقاء المواقع والمساحات الانتشارية معاً
            reports = reports.exclude(
                Q(location_type__name__icontains='دورة') |
                Q(location_type__name__icontains='مياه') |
                Q(location_type__name__icontains='حمام')
            )
            if category == 'location':
                reports = reports.filter(
                    Q(location_type__name__icontains='موقع') |
                    Q(location_type__name__icontains='مواقف') |
                    Q(location_type__name__icontains='مبنى')
                )
            elif category == 'open_space':
                reports = reports.filter(
                    Q(location_type__name__icontains='انتشار') |
                    Q(location_type__name__icontains='ساحة') |
                    Q(location_type__name__icontains='حديقة') |
                    Q(location_type__name__icontains='شارع') |
                    Q(location_type__name__icontains='طريق')
                )

        # التصدير المخصص لدورات المياه فقط
        if category == 'restroom':
            ws.title = 'تقارير دورات المياه'
            restroom_grouped = build_restroom_grouped_data(reports)
            max_visits = max((len(r['visits']) for r in restroom_grouped), default=1)

            headers = ['التاريخ', 'وحدة التنسيق', 'اسم دورة المياه']
            for v_idx in range(1, max_visits + 1):
                headers.extend([
                    f'زيارة {v_idx} - الموظف',
                    f'زيارة {v_idx} - الوقت',
                    f'زيارة {v_idx} - أدوات النظافة',
                    f'زيارة {v_idx} - مواد النظافة',
                    f'زيارة {v_idx} - أعمال النظافة الدورية لدورات المياه',
                    f'زيارة {v_idx} - الاعمال المدنية',
                    f'زيارة {v_idx} - الأعمال الكهربائية',
                    f'زيارة {v_idx} - الأعمال الميكانيكية'
                ])

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            ws.row_dimensions[1].height = 32

            for row_idx, r_data in enumerate(restroom_grouped, 2):
                row_vals = [r_data['date_str'], r_data['unit_name'], r_data['location_name']]
                for v in r_data['visits']:
                    row_vals.extend([v['user_name'], v['time_str']] + v['items'])

                missing_visits = max_visits - len(r_data['visits'])
                for _ in range(missing_visits):
                    row_vals.extend(['-', '-'] + [0]*6)

                for col, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.alignment = center
                    cell.border = border
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill('solid', fgColor='EBF5FB')

                ws.row_dimensions[row_idx].height = 25

            col_widths = [14, 20, 25] + [18, 14, 14, 14, 20, 14, 14, 14] * max_visits
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="restroom_reports_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            wb.save(response)
            return response

        # التصدير القياسي للمواقع الأخرى
        ws.title = 'التقارير الميدانية'
        headers = [
            'رقم التقرير', 'الموظف', 'وحدة التنسيق',
            'تصنيف الموقع', 'اسم الموقع', 'إحداثيات GPS', 'الملاحظات والأعمال', 'عدد الصور', 'تاريخ الإنشاء'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        ws.row_dimensions[1].height = 30

        reports = Report.objects.select_related(
            'user', 'unit', 'location_type', 'location'
        ).prefetch_related('items__note', 'items__work_type', 'items__images').order_by('-created_at')

        role = request.GET.get('role', 'contractor')
        category = request.GET.get('category')
        if role == 'supervisor':
            reports = reports.filter(user__profile__role='supervisor')
        elif role == 'contractor':
            reports = reports.filter(Q(user__profile__role='contractor') | Q(user__profile__isnull=True))

        if category == 'restroom':
            reports = reports.filter(
                Q(location_type__name__icontains='دورة') |
                Q(location_type__name__icontains='مياه') |
                Q(location_type__name__icontains='حمام')
            )
        elif category == 'location':
            reports = reports.filter(
                Q(location_type__name__icontains='موقع') |
                Q(location_type__name__icontains='مواقف') |
                Q(location_type__name__icontains='مبنى')
            )
        elif category == 'open_space':
            reports = reports.filter(
                Q(location_type__name__icontains='انتشار') |
                Q(location_type__name__icontains='ساحة') |
                Q(location_type__name__icontains='حديقة') |
                Q(location_type__name__icontains='شارع') |
                Q(location_type__name__icontains='طريق')
            )

        for row_num, report in enumerate(reports, 2):
            items_desc = []
        for row_num, report in enumerate(reports, 2):
            items_desc = []
            for item in report.items.all():
                wt = item.work_type.name if item.work_type else 'عمل ميداني'
                elem = f" - العنصر: {item.work_item_detail.name}" if item.work_item_detail and item.work_item_detail.name != 'عام' else ""
                act = item.work_activity.name if item.work_activity else item.work_classification
                act_desc = f" - البند: {act}" if act else ""
                items_desc.append(f"{wt}{elem}{act_desc}")

            
            notes_str = ' | '.join(items_desc)
            gps_str = f"{report.latitude}, {report.longitude}" if report.has_gps else 'غير مسجل'
            total_images = sum(item.images.count() for item in report.items.all()) + (1 if report.location_photo else 0)

            row_data = [
                report.report_number,
                report.user.get_full_name() or report.user.username,
                report.unit.name,
                report.location_type.name,
                report.location.name,
                gps_str,
                notes_str,
                total_images,
                report.created_at.strftime('%Y-%m-%d %H:%M'),
            ]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                cell.border = border
                if row_num % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='EBF5FB')

            ws.row_dimensions[row_num].height = 30

        col_widths = [18, 20, 22, 18, 22, 25, 50, 12, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reports_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response


# ─── تصدير PDF ───────────────────────────────────────────────────────────────
class ExportPDFView(StaffRequiredMixin, View):
    def get(self, request, pk):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER
        import os

        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            has_bidi = True
        except ImportError:
            has_bidi = False

        def ar(text):
            if not text:
                return ""
            text_str = str(text)
            if has_bidi:
                return get_display(arabic_reshaper.reshape(text_str))
            return text_str

        # Register Arabic Font
        font_name = 'Helvetica'
        for candidate in ['C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/tahoma.ttf', 'C:/Windows/Fonts/segoeui.ttf']:
            if os.path.exists(candidate):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', candidate))
                    font_name = 'ArabicFont'
                    break
                except Exception:
                    pass

        report = get_object_or_404(Report, pk=pk)
        items = list(report.items.select_related('note', 'work_type', 'work_item_detail', 'work_activity').prefetch_related('images'))

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),  # 29.7cm x 21.0cm
            rightMargin=1.5*cm, leftMargin=1.5*cm,
            topMargin=1.5*cm, bottomMargin=1.5*cm
        )

        usable_width = 26.7 * cm

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'RTLTitle', parent=styles['Title'],
            fontName=font_name,
            alignment=TA_CENTER, fontSize=22, spaceAfter=20,
            textColor=colors.HexColor('#1A5276')
        )
        section_title = ParagraphStyle(
            'RTLSection', parent=styles['Normal'],
            fontName=font_name,
            alignment=TA_CENTER, fontSize=16, spaceAfter=15,
            textColor=colors.HexColor('#1A5276')
        )
        normal_arabic = ParagraphStyle(
            'RTLNormal', parent=styles['Normal'],
            fontName=font_name,
            alignment=TA_RIGHT, fontSize=11,
            textColor=colors.HexColor('#2C3E50')
        )

        story = []

        # ── PAGE 1: المعلومات العامة للتقرير ──────────────────────────────────
        story.append(Paragraph(ar('تقرير ميداني شامل - بيانات التقرير والمعلومات العامة'), title_style))
        story.append(Spacer(1, 0.5*cm))

        gps_text = f"{report.latitude}, {report.longitude}" if report.has_gps else ar('غير مسجل')
        unit_name = report.unit.name if report.unit else 'غير محدد'
        loc_type_name = report.location_type.name if report.location_type else 'غير محدد'
        loc_name = report.location.name if report.location else 'غير محدد'
        user_name = report.user.get_full_name() or report.user.username if report.user else 'غير محدد'

        info_data = [
            [ar('رقم التقرير'), report.report_number],
            [ar('الموظف الميداني'), ar(user_name)],
            [ar('وحدة التنسيق'), ar(unit_name)],
            [ar('تصنيف الموقع'), ar(loc_type_name)],
            [ar('اسم الموقع'), ar(loc_name)],
            [ar('إحداثيات الموقع (GPS)'), gps_text],
            [ar('تاريخ الإنشاء'), report.created_at.strftime('%Y-%m-%d %H:%M')],
            [ar('إجمالي الملاحظات الميدانية'), str(len(items))],
        ]

        info_table = Table(info_data, colWidths=[7*cm, usable_width - 7*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#1A5276')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
            ('ROWBACKGROUNDS', (1, 0), (1, -1), [colors.white, colors.HexColor('#F4F6F7')]),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(info_table)

        # ── PAGE 2: الصورة العامة لتحديد الموقع ─────────────────────────────
        if report.location_photo and os.path.exists(report.location_photo.path):
            story.append(PageBreak())
            story.append(Paragraph(ar('الصورة العامة لتحديد الموقع'), section_title))
            story.append(Spacer(1, 0.4*cm))
            try:
                # Big image in landscape page
                loc_img = RLImage(report.location_photo.path, width=22*cm, height=13*cm)
                img_wrapper = Table([[loc_img]], colWidths=[usable_width])
                img_wrapper.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(img_wrapper)
            except Exception:
                pass

        # ── PAGES 3+: صفحة لكل ملاحظة مع صورها ────────────────────────────────
        for item_num, item in enumerate(items, 1):
            story.append(PageBreak())

            wt_name = item.work_type.name if item.work_type else 'غير محدد'
            elem_name = item.work_item_detail.name if item.work_item_detail and item.work_item_detail.name != 'عام' else ''
            act_name = item.work_activity.name if item.work_activity else item.work_classification or 'لا يوجد'
            note_text = item.note.text if item.note else ''

            # Banner Header for Item
            if elem_name:
                header_str = ar(f"الملاحظة رقم ({item_num} من {len(items)}) | نوع العمل: {wt_name} | العنصر: {elem_name} | البند: {act_name}")
            else:
                header_str = ar(f"الملاحظة رقم ({item_num} من {len(items)}) | نوع العمل: {wt_name} | البند: {act_name}")

            banner_data = [[header_str]]
            banner_table = Table(banner_data, colWidths=[usable_width])
            banner_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1A5276')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(banner_table)
            story.append(Spacer(1, 0.4*cm))

            if note_text:
                story.append(Paragraph(ar(f"تفاصيل الملاحظة: {note_text}"), normal_arabic))
                story.append(Spacer(1, 0.4*cm))

            images = list(item.images.all())
            if images:
                valid_paths = [img_obj.image.path for img_obj in images if os.path.exists(img_obj.image.path)]
                num_imgs = len(valid_paths)

                if num_imgs <= 2:
                    img_w, img_h = 12 * cm, 7.5 * cm
                    per_row = 2
                    col_w = 13 * cm
                elif num_imgs <= 4:
                    img_w, img_h = 10.5 * cm, 5.2 * cm
                    per_row = 2
                    col_w = 13 * cm
                else:
                    img_w, img_h = 7.5 * cm, 4.5 * cm
                    per_row = 3
                    col_w = 8.5 * cm

                valid_imgs = [RLImage(p, width=img_w, height=img_h) for p in valid_paths]

                if valid_imgs:
                    for i in range(0, len(valid_imgs), per_row):
                        row = valid_imgs[i:i+per_row]
                        while len(row) < per_row:
                            row.append('')
                        grid_table = Table([row], colWidths=[col_w] * per_row)
                        grid_table.setStyle(TableStyle([
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
                            ('TOPPADDING', (0, 0), (-1, -1), 3),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        ]))
                        story.append(grid_table)
                        story.append(Spacer(1, 0.25*cm))

        doc.build(story)
        buffer.seek(0)

        filename = f"{report.report_number}.pdf"
        as_attachment = request.GET.get('download') == '1'

        from django.http import FileResponse
        response = FileResponse(
            buffer,
            as_attachment=as_attachment,
            filename=filename,
            content_type='application/pdf'
        )
        if not as_attachment:
            response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response


class BulkExportPDFView(View):
    def post(self, request, *args, **kwargs):
        return self._generate_pdf(request)

    def get(self, request, *args, **kwargs):
        return self._generate_pdf(request)

    def _generate_pdf(self, request):
        ids_raw = request.POST.getlist('report_ids') or request.GET.getlist('report_ids')
        report_ids_str = request.POST.get('report_ids_str') or request.GET.get('report_ids_str')
        if not ids_raw and report_ids_str:
            ids_raw = report_ids_str.split(',')

        report_ids = []
        for rid in ids_raw:
            if isinstance(rid, str) and ',' in rid:
                report_ids.extend([int(x.strip()) for x in rid.split(',') if x.strip().isdigit()])
            elif str(rid).isdigit():
                report_ids.append(int(rid))

        if not report_ids:
            messages.error(request, 'لم يتم تحديد أي تقرير للتصدير.')
            return redirect(request.META.get('HTTP_REFERER', 'dashboard:report_list'))

        reports = Report.objects.filter(id__in=report_ids).select_related(
            'unit', 'location_type', 'location', 'user'
        ).prefetch_related(
            'items__note', 'items__work_type', 'items__work_item_detail', 'items__work_activity', 'items__images'
        ).order_by('-created_at')

        if not reports.exists():
            messages.error(request, 'التقارير المحددة غير موجودة.')
            return redirect(request.META.get('HTTP_REFERER', 'dashboard:report_list'))

        import io
        import os
        from django.utils import timezone
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER

        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            has_bidi = True
        except ImportError:
            has_bidi = False

        def ar(text):
            if not text:
                return ""
            text_str = str(text)
            if has_bidi:
                return get_display(arabic_reshaper.reshape(text_str))
            return text_str

        font_name = 'Helvetica'
        for candidate in ['C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/tahoma.ttf', 'C:/Windows/Fonts/segoeui.ttf']:
            if os.path.exists(candidate):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFontBulk', candidate))
                    font_name = 'ArabicFontBulk'
                    break
                except Exception:
                    pass

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.5*cm, leftMargin=1.5*cm,
            topMargin=1.5*cm, bottomMargin=1.5*cm
        )
        usable_width = 26.7 * cm

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'RTLTitle', parent=styles['Title'],
            fontName=font_name,
            alignment=TA_CENTER, fontSize=22, spaceAfter=20,
            textColor=colors.HexColor('#1A5276')
        )
        section_title = ParagraphStyle(
            'RTLSection', parent=styles['Normal'],
            fontName=font_name,
            alignment=TA_CENTER, fontSize=16, spaceAfter=15,
            textColor=colors.HexColor('#1A5276')
        )
        normal_arabic = ParagraphStyle(
            'RTLNormal', parent=styles['Normal'],
            fontName=font_name,
            alignment=TA_RIGHT, fontSize=11,
            textColor=colors.HexColor('#2C3E50')
        )

        story = []

        for idx, report in enumerate(reports):
            if idx > 0:
                story.append(PageBreak())

            items = list(report.items.all())

            # ── PAGE 1: المعلومات العامة للتقرير ──────────────────────────────────
            story.append(Paragraph(ar(f'تقرير ميداني شامل - {report.report_number}'), title_style))
            story.append(Spacer(1, 0.5*cm))

            gps_text = f"{report.latitude}, {report.longitude}" if report.has_gps else ar('غير مسجل')
            unit_name = report.unit.name if report.unit else 'غير محدد'
            loc_type_name = report.location_type.name if report.location_type else 'غير محدد'
            loc_name = report.location.name if report.location else 'غير محدد'
            user_name = report.user.get_full_name() or report.user.username if report.user else 'غير محدد'

            info_data = [
                [ar('رقم التقرير'), report.report_number],
                [ar('الموظف الميداني'), ar(user_name)],
                [ar('وحدة التنسيق'), ar(unit_name)],
                [ar('تصنيف الموقع'), ar(loc_type_name)],
                [ar('اسم الموقع'), ar(loc_name)],
                [ar('إحداثيات الموقع (GPS)'), gps_text],
                [ar('تاريخ الإنشاء'), report.created_at.strftime('%Y-%m-%d %H:%M')],
                [ar('إجمالي الملاحظات الميدانية'), str(len(items))],
            ]

            info_table = Table(info_data, colWidths=[7*cm, usable_width - 7*cm])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#1A5276')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
                ('ROWBACKGROUNDS', (1, 0), (1, -1), [colors.white, colors.HexColor('#F4F6F7')]),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(info_table)

            # ── PAGE 2: الصورة العامة لتحديد الموقع ─────────────────────────────
            if report.location_photo and os.path.exists(report.location_photo.path):
                story.append(PageBreak())
                story.append(Paragraph(ar('الصورة العامة لتحديد الموقع'), section_title))
                story.append(Spacer(1, 0.4*cm))
                try:
                    loc_img = RLImage(report.location_photo.path, width=22*cm, height=13*cm)
                    img_wrapper = Table([[loc_img]], colWidths=[usable_width])
                    img_wrapper.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]))
                    story.append(img_wrapper)
                except Exception:
                    pass

            # ── PAGES 3+: ملاحظات التقرير ──────────────────────────────────
            for item_num, item in enumerate(items, 1):
                story.append(PageBreak())

                wt_name = item.work_type.name if item.work_type else 'غير محدد'
                elem_name = item.work_item_detail.name if item.work_item_detail and item.work_item_detail.name != 'عام' else ''
                act_name = item.work_activity.name if item.work_activity else item.work_classification or 'لا يوجد'
                note_text = item.note.text if item.note else ''

                if elem_name:
                    header_str = ar(f"الملاحظة رقم ({item_num} من {len(items)}) | نوع العمل: {wt_name} | العنصر: {elem_name} | البند: {act_name}")
                else:
                    header_str = ar(f"الملاحظة رقم ({item_num} من {len(items)}) | نوع العمل: {wt_name} | البند: {act_name}")

                banner_table = Table([[header_str]], colWidths=[usable_width])
                banner_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1A5276')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                    ('FONTSIZE', (0, 0), (-1, -1), 12),
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))
                story.append(banner_table)
                story.append(Spacer(1, 0.4*cm))

                if note_text:
                    story.append(Paragraph(ar(f"تفاصيل الملاحظة: {note_text}"), normal_arabic))
                    story.append(Spacer(1, 0.4*cm))

                images = list(item.images.all())
                if images:
                    valid_paths = [img_obj.image.path for img_obj in images if os.path.exists(img_obj.image.path)]
                    num_imgs = len(valid_paths)

                    if num_imgs <= 2:
                        img_w, img_h = 12 * cm, 7.5 * cm
                        per_row = 2
                        col_w = 13 * cm
                    elif num_imgs <= 4:
                        img_w, img_h = 10.5 * cm, 5.2 * cm
                        per_row = 2
                        col_w = 13 * cm
                    else:
                        img_w, img_h = 7.5 * cm, 4.5 * cm
                        per_row = 3
                        col_w = 8.5 * cm

                    valid_imgs = [RLImage(p, width=img_w, height=img_h) for p in valid_paths]

                    if valid_imgs:
                        for i in range(0, len(valid_imgs), per_row):
                            row = valid_imgs[i:i+per_row]
                            while len(row) < per_row:
                                row.append('')
                            grid_table = Table([row], colWidths=[col_w] * per_row)
                            grid_table.setStyle(TableStyle([
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
                                ('TOPPADDING', (0, 0), (-1, -1), 3),
                                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                            ]))
                            story.append(grid_table)
                            story.append(Spacer(1, 0.25*cm))

        doc.build(story)
        buffer.seek(0)

        filename = f"merged_reports_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        from django.http import FileResponse
        return FileResponse(
            buffer,
            as_attachment=True,
            filename=filename,
            content_type='application/pdf'
        )
